import configparser
import csv
import fnmatch
import logging
import os
import sys
from collections import OrderedDict
from datetime import date, datetime
from math import isclose

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import quote_sheetname

from .process.cleansers import Cleanser

logger = logging.getLogger("bcompiler.utils")

rdel_cdel_merge = ""

DOCS = os.path.join(os.path.expanduser("~"), "Documents")
BCOMPILER_WORKING_D = "bcompiler"
ROOT_PATH = os.path.join(DOCS, BCOMPILER_WORKING_D)
SOURCE_DIR = os.path.join(ROOT_PATH, "source")

CONFIG_FILE = os.path.join(SOURCE_DIR, "config.ini")

runtime_config = configparser.ConfigParser()
runtime_config.read(CONFIG_FILE)

CURRENT_QUARTER = runtime_config["QuarterData"]["CurrentQuarter"]

try:
    SHEETS = [
        i for i in dict((runtime_config.items("TemplateSheets"))).values()
    ]
    BLANK_TEMPLATE_FN = runtime_config["BlankTemplate"]["name"]
except configparser.NoSectionError:
    print(
        "There is no config file present. Please run bcompiler-init to initialise bcompiler"
    )
    sys.exit()


def directory_has_returns_check(dir: str):
    if os.listdir(dir) == []:
        logger.critical(
            "Please copy populated return files to returns directory.")
        return False
    else:
        return True


def row_check(excel_file: str):
    wb = load_workbook(excel_file)
    data = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.rows
        data.append(
            dict(
                workbook=excel_file.split("/")[-1],
                sheet=sheet,
                row_count=len(list(rows)),
            ))
    return data


def row_data_formatter(csv_output=False, quiet=False) -> None:
    """
    Prints counts of rows in each sheet in each return spreadsheet.
    :param: csv_output - provide True to write output to csv file in output
    directory.
    :param: quiet - output differing row counts only. Cannot be used with
    csv_output argument.
    """
    if csv_output and quiet:
        logger.critical("Cannot use --csv and --quiet option. Choose one"
                        " or the other.")
        return
    try:
        returns_dir = os.path.join(ROOT_PATH, "source", "returns")
    except FileNotFoundError:
        logger.warning("There is no output directory. Run bcompiler -d to "
                       "set up working directories")
    try:
        tmpl_data = row_check(
            os.path.join(ROOT_PATH, "source", BLANK_TEMPLATE_FN))
    except FileNotFoundError:
        logger.warning("bicc_template.xlsm not found")
    if csv_output:
        csv_output_path = os.path.join(OUTPUT_DIR, "row_count.csv")
        csv_output_file = open(csv_output_path, "w", newline="")
        csv_writer = csv.writer(csv_output_file)
        logger.info("Writing output to csv file...")
    elif quiet:
        logger.info("Looking for anomolies in row counts in each sheet...")
    else:
        print("{0:<90}{1:<40}{2:<10}".format("Workbook", "Sheet", "Row Count"))
        print("{:#<150}".format(""))

    # Start with the bicc_template.xlsm BASE data
    for line in tmpl_data:
        if csv_output:
            csv_writer.writerow(
                [line["workbook"], line["sheet"], line["row_count"]])
        elif quiet:
            pass
        else:
            print(
                f"{line['workbook']:<90}{line['sheet']:<40}{line['row_count']:<10}"
            )
    print("{:#<150}".format(""))
    for f in os.listdir(returns_dir):
        if fnmatch.fnmatch(f, "*.xlsm"):
            d = row_check(os.path.join(returns_dir, f))
            zipped_data = zip(tmpl_data, d)
            for line in zipped_data:
                counts = [i["row_count"] for i in line]
                flag = counts[0] != counts[-1]
                if not flag:
                    if csv_output:
                        csv_writer.writerow([
                            line[1]["workbook"],
                            line[1]["sheet"],
                            line[1]["row_count"],
                        ])
                    elif quiet:
                        pass
                    else:
                        print(
                            f"{line[1]['workbook']:<90}{line[1]['sheet']:<40}{line[1]['row_count']:<10}"
                        )
                else:
                    if csv_output:
                        csv_writer.writerow([
                            line[1]["workbook"],
                            line[1]["sheet"],
                            line[1]["row_count"],
                            "INCONSISTENT WITH bicc_template.xlsm",
                        ])
                    else:
                        print(
                            f"{line[1]['workbook']:<90}{line[1]['sheet']:<40}{line[1]['row_count']:<10} *"
                        )
            if not quiet:
                print("{:#<150}".format(""))
            else:
                print(".")
        else:
            logger.critical(f"{f} does not have .xlsm file extension.")
    if csv_output:
        print(f"csv output file available at {csv_output_path}")
        csv_output_file.close()


def quick_typechecker(*args):
    """
    Very simple function to filter allowed types (int, float). Any other type
    returns False. All arguments must be of same type.
    """
    for arg in args:
        if isinstance(arg, (int, float, date)):
            pass
        else:
            return False
    return True


def simple_round(fl, prec):
    """Rounds a fl to prec precision."""
    return round(fl, prec)


def bc_is_close(x, y):
    """Returns true if acceptably close."""
    if isinstance(x, (date, datetime)) or isinstance(y, (date, datetime)):
        return False
    else:
        return isclose(x, y, rel_tol=0.001)


def cell_bg_colour(rgb=[]):
    """
    Give it a list of integers between 0 and 255 - three of them.
    """
    c_value = "{0:02X}{1:02X}{2:02X}".format(*rgb)
    return PatternFill(patternType="solid", fgColor=c_value, bgColor=c_value)


def get_relevant_names(project_name, project_data):

    try:
        sro_first_name = project_data[project_name]["SRO Full Name"].split(
            " ")[0]
    except IndexError:
        logger.warning(
            "SRO Full Name ({0}) is not suitable for splitting".format(
                project_data[project_name]["SRO Full Name"]))

    try:
        sro_last_name = project_data[project_name]["SRO Full Name"].split(
            " ")[1]
    except IndexError:
        logger.warning(
            "SRO Full Name ({0}) is not suitable for splitting".format(
                project_data[project_name]["SRO Full Name"]))

    try:
        pd_first_name = project_data[project_name]["PD Full Name"].split(
            " ")[0]
    except IndexError:
        logger.warning(
            "PD Full Name ({0}) is not suitable for splitting".format(
                project_data[project_name]["PD Full Name"]))

    try:
        pd_last_name = project_data[project_name]["PD Full Name"].split(" ")[1]
    except IndexError:
        logger.warning(
            "PD Full Name ({0}) is not suitable for splitting".format(
                project_data[project_name]["PD Full Name"]))

    try:
        sro_d = dict(first_name=sro_first_name, last_name=sro_last_name)
    except UnboundLocalError:
        sro_d = None
    try:
        pd_d = dict(first_name=pd_first_name, last_name=pd_last_name)
    except UnboundLocalError:
        pd_d = None

    return (sro_d, pd_d)


def project_data_from_master(master_file: str, opened_wb=False):
    if opened_wb is False:
        wb = load_workbook(master_file)
        ws = wb.active
    else:
        wb = master_file
        ws = wb.active
    # cleanse the keys
    for cell in ws["A"]:
        # we don't want to clean None...
        if cell.value is None:
            continue
        c = Cleanser(cell.value)
        cell.value = c.clean()
    p_dict = {}
    for col in ws.iter_cols(min_col=2):
        project_name = ""
        o = OrderedDict()
        for cell in col:
            if cell.row == 1:
                project_name = cell.value
                p_dict[project_name] = o
            else:
                val = ws.cell(row=cell.row, column=1).value
                if type(cell.value) == datetime:
                    d_value = date(cell.value.year, cell.value.month,
                                   cell.value.day)
                    p_dict[project_name][val] = d_value
                else:
                    p_dict[project_name][val] = cell.value
    # remove any "None" projects that were pulled from the master
    try:
        del p_dict[None]
    except KeyError:
        pass
    return p_dict


def project_data_line():
    p_dict = {}
    with open(SOURCE_DIR + "master_transposed.csv", "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = row.pop("Project/Programme Name")
            if key in p_dict:
                pass
            p_dict[key] = row
            logger.debug(
                "Adding {} to project_data_line dictionary".format(key))
    return p_dict


def open_openpyxl_template(template_file):
    """
    Opens an xlsx file (the template) and returns the openpyxl object.
    """
    wb = load_workbook(template_file, keep_vba=True)
    logger.info("Opening {} as an openpyxl object".format(template_file))
    return wb


def working_directory(dir_type=None):
    """
    Returns the working directory for source files
    :return: path to the working directory intended for the source files
    """
    docs = os.path.join(os.path.expanduser("~"), "Documents")
    bcomp_working_d = "bcompiler"
    try:
        root_path = os.path.join(docs, bcomp_working_d)
    except FileNotFoundError:
        print("You need to run with --create-wd to",
              "create the working directory")
    if dir_type == "source":
        return root_path + "/source/"
    elif dir_type == "output":
        return root_path + "/output/"
    elif dir_type == "returns":
        return root_path + "/source/returns/"
    else:
        return


# TODO this lot needs cleaning up - no more use of working_directory()

SOURCE_DIR = working_directory("source")
OUTPUT_DIR = working_directory("output")
RETURNS_DIR = working_directory("returns")
DATAMAP_RETURN_TO_MASTER = SOURCE_DIR + "datamap.csv"
DATAMAP_MASTER_TO_RETURN = SOURCE_DIR + "datamap.csv"
DATAMAP_MASTER_TO_GMPP = SOURCE_DIR + "archive/datamap-master-to-gmpp"
CLEANED_DATAMAP = SOURCE_DIR + "cleaned_datamap.csv"
MASTER = SOURCE_DIR + "master.csv"
TEMPLATE = SOURCE_DIR + BLANK_TEMPLATE_FN
GMPP_TEMPLATE = SOURCE_DIR + "archive/gmpp_template.xlsx"

