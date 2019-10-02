from collections import OrderedDict
from datetime import date
from datetime import datetime

from openpyxl import load_workbook

from datamaps.process import Cleanser


def project_data_from_master(master_file: str, opened_wb=False):
    if opened_wb is False:
        wb = load_workbook(master_file)
        ws = wb.active
    else:
        wb = master_file
        ws = wb.active
    # cleanse the keys
    for cell in ws["A"]:
        # we don't want to clean None...
        if cell.value is None:
            continue
        c = Cleanser(cell.value)
        cell.value = c.clean()
    p_dict = {}
    for col in ws.iter_cols(min_col=2):
        project_name = ""
        o = OrderedDict()
        for cell in col:
            if cell.row == 1:
                project_name = cell.value
                p_dict[project_name] = o
            else:
                val = ws.cell(row=cell.row, column=1).value
                if type(cell.value) == datetime:
                    d_value = date(cell.value.year, cell.value.month,
                                   cell.value.day)
                    p_dict[project_name][val] = d_value
                else:
                    p_dict[project_name][val] = cell.value
    # remove any "None" projects that were pulled from the master
    try:
        del p_dict[None]
    except KeyError:
        pass
    return p_dict